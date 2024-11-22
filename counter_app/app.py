from flask import Flask, render_template, request, redirect, url_for
import openpyxl
import os
from datetime import datetime

app = Flask(__name__)

# Get the current date in the format YYYY-MM-DD
current_date = datetime.now().strftime('%Y-%m-%d')

# Path where the Excel file will be saved with current date in the filename
EXCEL_FILE_PATH = os.path.join(os.getcwd(), 'data', f'counter_data_{current_date}.xlsx')

# Initialize count (it could also be stored in a session if you prefer)
counter = 0

# Ensure the directory exists
if not os.path.exists(os.path.dirname(EXCEL_FILE_PATH)):
    os.makedirs(os.path.dirname(EXCEL_FILE_PATH))

# Helper function to write count to Excel
def write_to_excel(count):
    # Check if the Excel file already exists
    if os.path.exists(EXCEL_FILE_PATH):
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    else:
        wb = openpyxl.Workbook()
    
    sheet = wb.active
    # Add headers if the file is empty
    if sheet.max_row == 1 and sheet.max_column == 1:
        sheet.append(['Timestamp', 'Count'])
    
    # Write the current count with a timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet.append([timestamp, count])
    wb.save(EXCEL_FILE_PATH)

@app.route('/')
def index():
    return render_template('index.html', counter=counter)

@app.route('/increase', methods=['POST'])
def increase():
    global counter
    counter += 1
    return redirect(url_for('index'))

@app.route('/decrease', methods=['POST'])
def decrease():
    global counter
    counter -= 1
    return redirect(url_for('index'))

@app.route('/submit', methods=['POST'])
def submit():
    global counter
    write_to_excel(counter)  # Save the current count to Excel
    return redirect(url_for('index'))

@app.route('/end_session', methods=['POST'])
def end_session():
    global counter
    counter = 0  # Reset the counter
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
